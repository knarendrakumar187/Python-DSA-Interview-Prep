"""Build a polished interview effort tracker Excel for Narendra (13-16 Aug)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import DoughnutChart, Reference
from pathlib import Path

wb = Workbook()

# ----- Theme -----
NAVY = "0F2C59"
TEAL = "1F6F5B"
ORANGE = "C65911"
GREEN = "006100"
RED = "9C0006"
GRAY = "333333"
LIGHT_NAVY = "D6E3F0"
LIGHT_TEAL = "E2F0EA"
LIGHT_ORANGE = "FCE4D6"
LIGHT_GREEN = "C6EFCE"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FFC7CE"
WHITE = "FFFFFF"
ROW_ALT = "F8FAFC"

thin = Border(
    left=Side(style="thin", color="D0D7E2"),
    right=Side(style="thin", color="D0D7E2"),
    top=Side(style="thin", color="D0D7E2"),
    bottom=Side(style="thin", color="D0D7E2"),
)
thick_bottom = Border(bottom=Side(style="medium", color=NAVY))

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_teal = PatternFill("solid", fgColor=TEAL)
fill_orange = PatternFill("solid", fgColor=LIGHT_ORANGE)
fill_green = PatternFill("solid", fgColor=LIGHT_GREEN)
fill_yellow = PatternFill("solid", fgColor=LIGHT_YELLOW)
fill_blue = PatternFill("solid", fgColor=LIGHT_NAVY)
fill_alt = PatternFill("solid", fgColor=ROW_ALT)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_card = PatternFill("solid", fgColor="EEF4FA")

font_title = Font(name="Calibri", bold=True, size=16, color=NAVY)
font_h = Font(name="Calibri", bold=True, size=11, color=WHITE)
font_sub = Font(name="Calibri", italic=True, size=10, color="555555")
font_body = Font(name="Calibri", size=10, color=GRAY)
font_bold = Font(name="Calibri", bold=True, size=11, color=NAVY)
font_kpi = Font(name="Calibri", bold=True, size=18, color=TEAL)
font_small = Font(name="Calibri", size=9, color="666666")

align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)


def width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers, fill=fill_navy):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = fill
        cell.font = font_h
        cell.alignment = align_c
        cell.border = thin


def style_cells(ws, r, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(r, c)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_c if c not in (4, 9, 3) else align_l
        if alt and cell.fill.fgColor is None or (alt and getattr(cell.fill.fgColor, "rgb", None) in (None, "00000000")):
            pass
    if alt:
        for c in range(1, cols + 1):
            if ws.cell(r, c).fill.fgColor is None or str(getattr(ws.cell(r, c).fill.fgColor, "rgb", "")) in ("00000000", "None"):
                ws.cell(r, c).fill = fill_alt


def dv_list(ws, ref, options):
    dv = DataValidation(type="list", formula1=options, allow_blank=True)
    dv.error = "Choose from dropdown"
    dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    dv.add(ref)


def banner(ws, title, subtitle, merge="A1:J1"):
    ws.merge_cells(merge)
    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 28
    a2, b2 = merge.split(":")[0][0] + "2", merge.split(":")[1][0] + "2"
    ws.merge_cells(f"A2:{merge.split(':')[1][0]}2")
    ws["A2"] = subtitle
    ws["A2"].font = font_sub
    ws.row_dimensions[2].height = 18


# ========================= 00 DASHBOARD =========================
ws = wb.active
ws.title = "00_Dashboard"
banner(
    ws,
    "NARENDRA KUMAR — INTERVIEW COMMAND CENTER",
    "Interview: 16 Aug 2026  |  Effort window: 13–15 Aug  |  Goal: Perform better with depth (not perfection)",
    "A1:H1",
)

# KPI cards labels
kpis = [
    (4, "A", "B", "Overall Done", "=COUNTIF('01_Master_Plan'!H5:H37,\"Done\")&\" / 33\""),
    (4, "C", "D", "DSA Done", "=COUNTIF('02_DSA_Problems'!H4:H15,\"Done\")&\" / 12\""),
    (4, "E", "F", "Resume Speak Done", "=COUNTIF('03_Resume_HR'!F4:F14,\"Done\")&\" / 11\""),
    (4, "G", "H", "CS Flash Done", "=COUNTIF('04_CS_Core'!F4:F17,\"Done\")&\" / 14\""),
]

# Manual KPI layout
ws["A4"] = "OVERALL PROGRESS"
ws["A4"].font = font_bold
ws.merge_cells("A4:B4")
ws["A5"] = '=COUNTIF(\'01_Master_Plan\'!H5:H37,"Done")'
ws["B5"] = '/ 33 tasks'
ws["A5"].font = font_kpi
ws["A6"] = '=IFERROR(ROUND(COUNTIF(\'01_Master_Plan\'!H5:H37,"Done")/33*100,0)&"% complete","0%")'
ws["A6"].font = Font(name="Calibri", bold=True, size=12, color=ORANGE)

ws["C4"] = "DSA PROBLEMS"
ws["C4"].font = font_bold
ws.merge_cells("C4:D4")
ws["C5"] = '=COUNTIF(\'02_DSA_Problems\'!H4:H15,"Done")'
ws["D5"] = "/ 12"
ws["C5"].font = font_kpi
ws["C6"] = '=IFERROR(ROUND(COUNTIF(\'02_DSA_Problems\'!H4:H15,"Done")/12*100,0)&"%","0%")'
ws["C6"].font = Font(name="Calibri", bold=True, size=12, color=ORANGE)

ws["E4"] = "RESUME / HR"
ws["E4"].font = font_bold
ws.merge_cells("E4:F4")
ws["E5"] = '=COUNTIF(\'03_Resume_HR\'!F4:F14,"Done")'
ws["F5"] = "/ 11"
ws["E5"].font = font_kpi
ws["E6"] = '=IFERROR(ROUND(COUNTIF(\'03_Resume_HR\'!F4:F14,"Done")/11*100,0)&"%","0%")'
ws["E6"].font = Font(name="Calibri", bold=True, size=12, color=ORANGE)

ws["G4"] = "CS CORE"
ws["G4"].font = font_bold
ws.merge_cells("G4:H4")
ws["G5"] = '=COUNTIF(\'04_CS_Core\'!F4:F17,"Done")'
ws["H5"] = "/ 14"
ws["G5"].font = font_kpi
ws["G6"] = '=IFERROR(ROUND(COUNTIF(\'04_CS_Core\'!F4:F17,"Done")/14*100,0)&"%","0%")'
ws["G6"].font = Font(name="Calibri", bold=True, size=12, color=ORANGE)

for col in ["A", "C", "E", "G"]:
    for r in range(4, 7):
        ws[f"{col}{r}"].fill = fill_card
        ws[f"{col}{r}"].border = thin
        ws[f"{chr(ord(col)+1)}{r}"].fill = fill_card
        ws[f"{chr(ord(col)+1)}{r}"].border = thin

ws["A8"] = "TODAY FOCUS (read this first)"
ws["A8"].font = font_bold
ws.merge_cells("A8:H8")

focus = [
    ["Date", "Priority Focus", "Hours", "Must Finish Before Sleep"],
    ["Thu 13 Aug (Tonight)", "4 DSA + Intro B + Nyay + PK/FK + Process/Thread", "2.5–3 hrs", "Two Sum, Duplicate, Stock, Parentheses + Intro spoken"],
    ["Fri 14 Aug (Main Day)", "Window/LL/BS/Tree + CS flash + full resume speak", "4–5 hrs", "Reverse LL, Binary Search, Tree depth + HR answers"],
    ["Sat 15 Aug (Polish)", "Fails only + Full mock + sleep early", "3–4 hrs", "Full mock done + no new topics"],
    ["Sun 16 Aug (Interview)", "Intro + Nyay only (15 min). Stay calm.", "light", "Clarify→Example→Brute→Better→Code→Complexity"],
]
header_row(ws, 9, focus[0], fill_teal)
for i, row in enumerate(focus[1:], 10):
    for c, val in enumerate(row, 1):
        cell = ws.cell(i, c, val)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_l if c > 1 else align_c
        if i % 2 == 0:
            cell.fill = fill_alt
    ws.row_dimensions[i].height = 34
ws.merge_cells("B10:B10")

ws["A15"] = "HOW TO USE (simple)"
ws["A15"].font = font_bold
tips = [
    "1. Work mainly in sheet 01_Master_Plan — mark Status = Done only if you can explain out loud.",
    "2. For every DSA problem: write Brute idea + Optimal idea, then code Optimal.",
    "3. Orange Priority = MUST. Do MUST before SHOULD.",
    "4. Use 03_Resume_HR for speaking practice. Record yourself once.",
    "5. Do not add random YouTube topics. This sheet is enough for a fresher technical round baseline.",
    "6. Target before interview morning: Overall >= 75% Done.",
]
for i, t in enumerate(tips, 16):
    ws.cell(i, 1, t).font = font_body
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

ws["A23"] = "FILE PATHS"
ws["A23"].font = font_bold
ws["A24"] = r"Local folder: E:\sems\sem 7\Python DSA"
ws["A25"] = r"Intro PDF: Resume_Prep\Narendra_Interview_Intro.pdf"
ws["A26"] = r"HR PDF: Resume_Prep\Narendra_Interview_Behavioural.pdf"
ws["A27"] = "GitHub: https://github.com/knarendrakumar187/Python-DSA-Interview-Prep"
for r in range(24, 28):
    ws.cell(r, 1).font = font_small
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

ws["A29"] = "MINDSET"
ws["A29"].font = font_bold
ws["A30"] = "You are not behind forever. You are 3 focused days away from a clearer interview. Depth > panic. Start Task #1 now."
ws["A30"].font = Font(name="Calibri", italic=True, size=11, color=TEAL)
ws.merge_cells("A30:H30")

width(ws, [22, 48, 14, 18, 14, 14, 14, 18])
ws.freeze_panes = "A4"

# ========================= 01 MASTER PLAN =========================
ws1 = wb.create_sheet("01_Master_Plan")
banner(
    ws1,
    "01 — MASTER PLAN (13 Aug night → 16 Aug)",
    "Mark Status with dropdown. MUST rows are critical. Progress auto-counts on Dashboard.",
    "A1:J1",
)

ws1["A3"] = "Filter by Date/Priority using the dropdown arrows in the header row."
ws1["A3"].font = font_small
ws1.merge_cells("A3:J3")

headers = ["#", "Date", "Block", "Task", "Category", "Priority", "Mins", "Status", "Resource / File", "My Notes"]
header_row(ws1, 4, headers)

tasks = [
    [1, "Thu 13 Aug", "Night-1", "Solve Two Sum (explain brute + code hash map)", "DSA", "MUST", 25, "Not Started", "Day01_02_Arrays/Day01_problems.py", ""],
    [2, "Thu 13 Aug", "Night-1", "Solve Contains Duplicate", "DSA", "MUST", 15, "Not Started", "Day01_problems.py", ""],
    [3, "Thu 13 Aug", "Night-1", "Solve Best Time to Buy and Sell Stock", "DSA", "MUST", 20, "Not Started", "Day01_problems.py", ""],
    [4, "Thu 13 Aug", "Night-1", "Solve Valid Parentheses (stack)", "DSA", "MUST", 20, "Not Started", "Day07_Stack_Queue/Day07_problems.py", ""],
    [5, "Thu 13 Aug", "Night-2", "Speak Intro Version B three times with timer", "Resume", "MUST", 15, "Not Started", "Resume_Prep/INTRO.md or PDF", ""],
    [6, "Thu 13 Aug", "Night-2", "Speak Nyay Sahayak: What + RAG + 1 difficulty", "Resume", "MUST", 20, "Not Started", "Resume_Prep/PROJECT_DEPTH.md", ""],
    [7, "Thu 13 Aug", "Night-2", "CS speak: Primary key vs Foreign key + JOIN types", "CS", "MUST", 20, "Not Started", "CS_Core/DBMS_SQL", ""],
    [8, "Thu 13 Aug", "Night-2", "CS speak: Process vs Thread with example", "CS", "MUST", 10, "Not Started", "CS_Core/OS", ""],
    [9, "Fri 14 Aug", "Morning", "Solve Move Zeroes (in-place two pointer idea)", "DSA", "MUST", 20, "Not Started", "Day01 / arrays", ""],
    [10, "Fri 14 Aug", "Morning", "Solve Valid Palindrome (two pointers)", "DSA", "MUST", 20, "Not Started", "Day04_Two_Pointers", ""],
    [11, "Fri 14 Aug", "Morning", "Attempt Longest Substring Without Repeating", "DSA", "MUST", 35, "Not Started", "Day05_Sliding_Window", ""],
    [12, "Fri 14 Aug", "Morning", "Reverse Linked List + draw prev/curr/next", "DSA", "MUST", 30, "Not Started", "Day08_LinkedList", ""],
    [13, "Fri 14 Aug", "Morning", "Binary Search template from memory", "DSA", "MUST", 25, "Not Started", "Day10_Sorting_BinarySearch", ""],
    [14, "Fri 14 Aug", "Afternoon", "Max Depth of Binary Tree", "DSA", "MUST", 25, "Not Started", "Day11_Trees", ""],
    [15, "Fri 14 Aug", "Afternoon", "Timed drill: 2 Easy problems in 40 minutes", "Mock", "MUST", 40, "Not Started", "Timer on phone", ""],
    [16, "Fri 14 Aug", "Afternoon", "Rewrite all FAIL problems once", "DSA", "MUST", 40, "Not Started", "List fails in My Notes", ""],
    [17, "Fri 14 Aug", "Evening", "Speak Intro Version A (under 2 minutes)", "Resume", "MUST", 15, "Not Started", "INTRO.md", ""],
    [18, "Fri 14 Aug", "Evening", "GeoVerse 1-min pitch + 1 difficulty", "Resume", "SHOULD", 15, "Not Started", "PROJECT_DEPTH.md", ""],
    [19, "Fri 14 Aug", "Evening", "AWS pipeline 1-min pitch + 1 difficulty", "Resume", "SHOULD", 15, "Not Started", "PROJECT_DEPTH.md", ""],
    [20, "Fri 14 Aug", "Evening", "HR: Strength + Weakness + Why hire you", "HR", "MUST", 20, "Not Started", "BEHAVIOURAL.md / PDF", ""],
    [21, "Fri 14 Aug", "Evening", "CS: ACID properties with UPI example", "CS", "MUST", 15, "Not Started", "DBMS_SQL/03", ""],
    [22, "Fri 14 Aug", "Evening", "CS: Deadlock 4 conditions + real example", "CS", "MUST", 15, "Not Started", "OS/02", ""],
    [23, "Fri 14 Aug", "Evening", "CS: TCP vs UDP + HTTP vs HTTPS + DNS", "CS", "MUST", 20, "Not Started", "CN notes", ""],
    [24, "Fri 14 Aug", "Evening", "OOP 4 pillars with tiny Python example", "CS", "SHOULD", 15, "Not Started", "CS_Core/OOP.md", ""],
    [25, "Sat 15 Aug", "Morning", "Re-solve ONLY failed DSA problems", "DSA", "MUST", 90, "Not Started", "No new topics", ""],
    [26, "Sat 15 Aug", "Afternoon", "FULL MOCK: Intro → 2 DSA → Nyay → CS → HR", "Mock", "MUST", 75, "Not Started", "Day14_FULL_MOCK_SCRIPT.md", ""],
    [27, "Sat 15 Aug", "Afternoon", "Write top 5 weak answers and fix same day", "Mock", "MUST", 45, "Not Started", "Notebook / Notes col", ""],
    [28, "Sat 15 Aug", "Evening", "Flash revise CHEATSHEET + CS QUICK_REVISE", "Revise", "MUST", 30, "Not Started", "CHEATSHEET.md", ""],
    [29, "Sat 15 Aug", "Evening", "Final speak: Intro + Nyay + Why hire you", "Resume", "MUST", 20, "Not Started", "", ""],
    [30, "Sat 15 Aug", "Night", "Sleep before 11 PM. No new coding.", "Health", "MUST", 0, "Not Started", "Critical for performance", ""],
    [31, "Sun 16 Aug", "Morning", "15 min only: Intro + Nyay Sahayak", "Interview", "MUST", 15, "Not Started", "No new DSA", ""],
    [32, "Sun 16 Aug", "Morning", "Water + resume ready + calm breathing", "Interview", "MUST", 10, "Not Started", "", ""],
    [33, "Sun 16 Aug", "Interview", "Use formula: Clarify→Example→Brute→Better→Code→Complexity", "Interview", "MUST", 0, "Not Started", "Keep in mind", ""],
]

for i, row in enumerate(tasks):
    r = 5 + i
    for c, val in enumerate(row, 1):
        cell = ws1.cell(r, c, val)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_l if c in (4, 9, 10) else align_c
        if i % 2 == 1:
            cell.fill = fill_alt
        if row[5] == "MUST" and c == 6:
            cell.fill = fill_orange
            cell.font = Font(name="Calibri", bold=True, size=10, color=ORANGE)
        if "16 Aug" in row[1] and c == 2:
            cell.fill = fill_blue
        if row[1] == "Sat 15 Aug" and row[2] == "Night" and c == 2:
            cell.fill = fill_yellow
    ws1.row_dimensions[r].height = 28

dv_list(ws1, "H5:H37", '"Not Started,In Progress,Done"')
dv_list(ws1, "F5:F37", '"MUST,SHOULD"')

# conditional formatting for Status Done
ws1.conditional_formatting.add(
    "H5:H37",
    FormulaRule(formula=['$H5="Done"'], fill=fill_green),
)
ws1.conditional_formatting.add(
    "H5:H37",
    FormulaRule(formula=['$H5="In Progress"'], fill=fill_yellow),
)

width(ws1, [4, 13, 11, 56, 10, 10, 7, 13, 36, 22])
ws1.freeze_panes = "A5"
ws1.auto_filter.ref = "A4:J37"

ws1["A39"] = "QUICK COUNTS"
ws1["A39"].font = font_bold
ws1["A40"] = "Done:"
ws1["B40"] = '=COUNTIF(H5:H37,"Done")'
ws1["C40"] = "In Progress:"
ws1["D40"] = '=COUNTIF(H5:H37,"In Progress")'
ws1["E40"] = "Not Started:"
ws1["F40"] = '=COUNTIF(H5:H37,"Not Started")'
ws1["A41"] = "MUST completed:"
ws1["B41"] = '=COUNTIFS(F5:F37,"MUST",H5:H37,"Done")'
ws1["C41"] = "MUST total:"
ws1["D41"] = '=COUNTIF(F5:F37,"MUST")'
ws1["E41"] = "% MUST done:"
ws1["F41"] = '=IFERROR(ROUND(B41/D41*100,0)&"%","0%")'
ws1["F41"].font = Font(name="Calibri", bold=True, size=12, color=TEAL)

# ========================= 02 DSA =========================
ws2 = wb.create_sheet("02_DSA_Problems")
banner(
    ws2,
    "02 — DSA MUST PROBLEMS (Brute + Optimal)",
    "Interview habit: explain brute first, then code optimal, then say Time & Space.",
    "A1:I1",
)
header_row(ws2, 3, ["#", "Problem", "Pattern", "Brute Idea", "Optimal Idea", "Coded?", "Explain T/S?", "Status", "Confidence (1-5)"])

dsa = [
    [1, "Two Sum", "Hash Map", "Check all pairs O(n²)", "Hash map value→index O(n)", "No", "No", "Not Started", ""],
    [2, "Contains Duplicate", "Hash Set", "Nested loops", "Use set / compare lengths", "No", "No", "Not Started", ""],
    [3, "Best Time to Buy/Sell Stock", "One Pass", "Try all buy/sell pairs", "Track min price so far", "No", "No", "Not Started", ""],
    [4, "Move Zeroes", "Two Pointers", "Build new array", "Write pointer in-place", "No", "No", "Not Started", ""],
    [5, "Valid Parentheses", "Stack", "Count only (incomplete)", "Stack + matching pairs", "No", "No", "Not Started", ""],
    [6, "Valid Palindrome", "Two Pointers", "Clean string + reverse", "L/R pointers skip junk", "No", "No", "Not Started", ""],
    [7, "Longest Substring No Repeat", "Sliding Window", "Generate all substrings", "Window + last-seen map", "No", "No", "Not Started", ""],
    [8, "Reverse Linked List", "Pointers", "Copy values to array", "prev / curr / next reverse", "No", "No", "Not Started", ""],
    [9, "Binary Search", "Divide & Conquer", "Linear scan O(n)", "Mid cut half O(log n)", "No", "No", "Not Started", ""],
    [10, "Max Depth of Binary Tree", "DFS / Recursion", "BFS level count also ok", "1 + max(left, right)", "No", "No", "Not Started", ""],
    [11, "Majority Element (bonus)", "Hashing", "Count with dict", "Counter / Boyer-Moore optional", "No", "No", "Not Started", ""],
    [12, "First Unique Character (bonus)", "Hashing", "Nested scan", "Counter then one pass", "No", "No", "Not Started", ""],
]

for i, row in enumerate(dsa):
    r = 4 + i
    for c, val in enumerate(row, 1):
        cell = ws2.cell(r, c, val)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_l if c in (2, 4, 5) else align_c
        if i % 2 == 1:
            cell.fill = fill_alt
    ws2.row_dimensions[r].height = 34

dv_list(ws2, "F4:F15", '"No,Yes"')
dv_list(ws2, "G4:G15", '"No,Yes"')
dv_list(ws2, "H4:H15", '"Not Started,In Progress,Done"')
dv_list(ws2, "I4:I15", '"1,2,3,4,5"')
ws2.conditional_formatting.add("H4:H15", FormulaRule(formula=['$H4="Done"'], fill=fill_green))
ws2.conditional_formatting.add("H4:H15", FormulaRule(formula=['$H4="In Progress"'], fill=fill_yellow))

width(ws2, [4, 30, 16, 28, 30, 10, 13, 13, 14])
ws2.freeze_panes = "A4"
ws2.auto_filter.ref = "A3:I15"

ws2["A17"] = "SPEAK TEMPLATE IN INTERVIEW"
ws2["A17"].font = font_bold
ws2["A18"] = "1) Clarify input/output  2) Dry-run example  3) Brute force + complexity  4) Optimal + complexity  5) Code  6) Edge cases"
ws2["A18"].font = Font(name="Calibri", size=11, color=TEAL)
ws2.merge_cells("A18:I18")
ws2["A20"] = "Done count:"
ws2["B20"] = '=COUNTIF(H4:H15,"Done")'
ws2["C20"] = "Avg confidence:"
ws2["D20"] = '=IFERROR(ROUND(AVERAGE(I4:I15),1),"-")'

# ========================= 03 RESUME =========================
ws3 = wb.create_sheet("03_Resume_HR")
banner(
    ws3,
    "03 — RESUME + HR SPEAKING TRACKER",
    "Mark Done only when you can speak without reading notes.",
    "A1:G1",
)
header_row(ws3, 3, ["#", "Item", "Must Cover", "Target Time", "Without Notes?", "Status", "Last Practiced"])

resume = [
    [1, "Intro Version B (short)", "Name, college, CGPA, AWS cert, 3 projects short, role goal", "60–75 sec", "No", "Not Started", ""],
    [2, "Intro Version A (full)", "Short family + projects + hackathons + photography + close", "90–110 sec", "No", "Not Started", ""],
    [3, "Nyay Sahayak deep dive", "RAG flow, ChromaDB+Groq, roles, IPC↔BNS, difficulty, next improvement", "3 min", "No", "Not Started", ""],
    [4, "What is RAG?", "Retrieve relevant docs, then generate; reduces hallucination", "45 sec", "No", "Not Started", ""],
    [5, "GeoVerse pitch", "Live APIs, AI itinerary, fallback, lazy routes/cache", "60 sec", "No", "Not Started", ""],
    [6, "AWS Intelligence Loop", "Lambda + Comprehend + DynamoDB + QuickSight + one difficulty", "60 sec", "No", "Not Started", ""],
    [7, "Strength", "Ship usable products (auth/fallback/RAG) + hackathon calm", "45 sec", "No", "Not Started", ""],
    [8, "Weakness", "Earlier surface-level; mock said needs depth; now practice depth daily", "45 sec", "No", "Not Started", ""],
    [9, "Why hire you?", "Learn fast + shipped projects + AWS cert + ownership mindset", "45 sec", "No", "Not Started", ""],
    [10, "STAR hardest bug", "Nyay ungrounded answers → improved retrieval/RAG → learning", "60–90 sec", "No", "Not Started", ""],
    [11, "3–5 year vision", "Own features end-to-end; later mentor juniors; now fundamentals", "40 sec", "No", "Not Started", ""],
]

for i, row in enumerate(resume):
    r = 4 + i
    for c, val in enumerate(row, 1):
        cell = ws3.cell(r, c, val)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_l if c in (2, 3) else align_c
        if i % 2 == 1:
            cell.fill = fill_alt
    ws3.row_dimensions[r].height = 40

dv_list(ws3, "E4:E14", '"No,Yes"')
dv_list(ws3, "F4:F14", '"Not Started,In Progress,Done"')
ws3.conditional_formatting.add("F4:F14", FormulaRule(formula=['$F4="Done"'], fill=fill_green))
width(ws3, [4, 28, 70, 12, 14, 13, 14])
ws3.freeze_panes = "A4"

# ========================= 04 CS =========================
ws4 = wb.create_sheet("04_CS_Core")
banner(
    ws4,
    "04 — CS CORE FLASH CARDS",
    "Answer style: Definition + small example + why it matters in real systems.",
    "A1:G1",
)
header_row(ws4, 3, ["#", "Subject", "Question", "Answer Keywords", "Can Speak?", "Status", "Revise Count"])

cs = [
    [1, "DBMS", "Primary key vs Foreign key", "PK uniquely identifies row; FK links to another table", "No", "Not Started", 0],
    [2, "DBMS", "INNER JOIN vs LEFT JOIN", "INNER=matches only; LEFT=all left rows + matches", "No", "Not Started", 0],
    [3, "DBMS", "ACID (with UPI example)", "Atomicity Consistency Isolation Durability", "No", "Not Started", 0],
    [4, "DBMS", "What is an Index?", "Faster reads; slightly slower writes", "No", "Not Started", 0],
    [5, "DBMS", "Normalization meaning", "Reduce duplication and update anomalies", "No", "Not Started", 0],
    [6, "OS", "Process vs Thread", "Process=own memory; Thread=shared memory workers", "No", "Not Started", 0],
    [7, "OS", "Deadlock 4 conditions", "Mutual exclusion, Hold&Wait, No preemption, Circular wait", "No", "Not Started", 0],
    [8, "OS", "Virtual memory / page fault", "Use disk as extra memory; fault loads missing page", "No", "Not Started", 0],
    [9, "CN", "TCP vs UDP", "TCP reliable ordered; UDP faster, less reliable", "No", "Not Started", 0],
    [10, "CN", "HTTP vs HTTPS", "HTTPS = HTTP + TLS encryption", "No", "Not Started", 0],
    [11, "CN", "DNS purpose", "Converts domain name to IP address", "No", "Not Started", 0],
    [12, "OOP", "Four pillars + example", "Encapsulation, Abstraction, Inheritance, Polymorphism", "No", "Not Started", 0],
    [13, "SE", "Agile vs Waterfall", "Agile=sprints+feedback; Waterfall=sequential stages", "No", "Not Started", 0],
    [14, "SQL", "Write JOIN + GROUP BY query", "Practice from CS_Core/DBMS_SQL/04_sql_practice.md", "No", "Not Started", 0],
]

for i, row in enumerate(cs):
    r = 4 + i
    for c, val in enumerate(row, 1):
        cell = ws4.cell(r, c, val)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_l if c in (3, 4) else align_c
        if i % 2 == 1:
            cell.fill = fill_alt
    ws4.row_dimensions[r].height = 32

dv_list(ws4, "E4:E17", '"No,Yes"')
dv_list(ws4, "F4:F17", '"Not Started,In Progress,Done"')
ws4.conditional_formatting.add("F4:F17", FormulaRule(formula=['$F4="Done"'], fill=fill_green))
width(ws4, [4, 10, 36, 58, 12, 13, 12])
ws4.freeze_panes = "A4"
ws4.auto_filter.ref = "A3:G17"

# ========================= 05 Interview day =========================
ws5 = wb.create_sheet("05_Interview_Day")
banner(
    ws5,
    "05 — 16 AUG INTERVIEW DAY",
    "Keep this sheet simple. No new learning. Execute calmly.",
    "A1:D1",
)
header_row(ws5, 3, ["#", "Checklist Item", "Status", "Notes"], fill_teal)

day = [
    [1, "Slept well the night before", "Not Started", ""],
    [2, "Light breakfast + water bottle ready", "Not Started", ""],
    [3, "Resume PDF / print ready", "Not Started", ""],
    [4, "Laptop + internet checked (if online interview)", "Not Started", ""],
    [5, "15 min revise ONLY: Intro + Nyay Sahayak", "Not Started", "No new coding"],
    [6, "Reach venue / join link 10 minutes early", "Not Started", ""],
    [7, "Greet interviewer calmly; smile once", "Not Started", ""],
    [8, "DSA formula ready in mind: Clarify→Example→Brute→Better→Code→T/S", "Not Started", ""],
    [9, "If stuck: explain brute force first, then optimize", "Not Started", ""],
    [10, "Be honest about any resume number you cannot defend", "Not Started", ""],
    [11, "Ask 1 good question at end (team stack / mentorship)", "Not Started", ""],
    [12, "After interview: write questions asked + mistakes", "Not Started", "Learn for next round"],
]

for i, row in enumerate(day):
    r = 4 + i
    for c, val in enumerate(row, 1):
        cell = ws5.cell(r, c, val)
        cell.border = thin
        cell.font = font_body
        cell.alignment = align_l if c in (2, 4) else align_c
        if i % 2 == 1:
            cell.fill = fill_alt
    ws5.row_dimensions[r].height = 26

dv_list(ws5, "C4:C15", '"Not Started,Done"')
ws5.conditional_formatting.add("C4:C15", FormulaRule(formula=['$C4="Done"'], fill=fill_green))
width(ws5, [5, 78, 14, 28])

ws5["A17"] = "FINAL LINE TO REMEMBER"
ws5["A17"].font = font_bold
ws5["A18"] = "You already built real projects. In this interview, show clear thinking. One round does not define your whole career."
ws5["A18"].font = Font(name="Calibri", italic=True, size=11, color=TEAL)
ws5.merge_cells("A18:D18")

# ========================= 06 Weak log =========================
ws6 = wb.create_sheet("06_Weak_Log")
banner(
    ws6,
    "06 — WEAK LOG (write fails here)",
    "Every fail becomes a revision item. Fix within 24 hours.",
    "A1:F1",
)
header_row(ws6, 3, ["#", "Date", "Topic / Problem", "What went wrong", "Fix / Correct idea", "Fixed?"], fill=PatternFill("solid", fgColor=ORANGE))

for i in range(15):
    r = 4 + i
    ws6.cell(r, 1, i + 1).border = thin
    for c in range(2, 7):
        cell = ws6.cell(r, c, "")
        cell.border = thin
        cell.alignment = align_l
        if i % 2 == 1:
            cell.fill = fill_alt
    ws6.row_dimensions[r].height = 30

dv_list(ws6, "F4:F18", '"No,Yes"')
width(ws6, [5, 14, 28, 40, 40, 10])

out = Path(r"E:\sems\sem 7\Python DSA\Narendra_Interview_Tracker_FINAL.xlsx")
wb.save(out)
print("Saved:", out)
print("Sheets:", wb.sheetnames)
